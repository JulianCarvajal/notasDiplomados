import pyodbc
import openpyxl
import os

server = '(localdb)\MSSQLLocalDB'
basesdatos= 'Dimplomados'
user = 'sa'
pasword = 'David3110'
xls=os.getcwd()
excel_document= openpyxl.load_workbook(xls+'\ReporteNotas_Ejemplo.xlsx')
sheet = excel_document['Hoja1']
lnMaxFilas = sheet.max_row

try:
    conexion = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; SERVER=' + server+';DATABASE='+basesdatos+ ';UID='+user+';PWD='+pasword)
    print('Conexion lista')

except:
    print('falla la conexion')

fila=1
for i in range(lnMaxFilas):
    tipoid=sheet.cell(fila,1).value
    if tipoid == None:
        break
    Cedula=sheet.cell(fila,2).value
    Cod_Est=sheet.cell(fila,3).value
    Nombres=sheet.cell(fila,4).value
    Apellidos=sheet.cell(fila,5).value
    email=sheet.cell(fila,7).value
    telefon=sheet.cell(fila,8).value
    genero=sheet.cell(fila,9).value
    Admin=sheet.cell(fila,10).value
    data=conexion.cursor()
    data.execute("INSERT INTO Usuarios (Cod_ID, Cedula, Cod_Usuario, Nombres, Apellidos, Cod_Dip, Email, Tel, Cod_Genero, Cod_Admin) VALUES ('{}', '{}','{}','{}','{}','PY', '{}', '{}', '{}', '{}')".format(tipoid, Cedula, Cod_Est, Nombres, Apellidos,email, telefon, genero, Admin))
    data.commit()
    fila=fila+1